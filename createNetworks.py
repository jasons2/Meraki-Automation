# -*- coding: utf-8 -*-

__author__ = "Jason Smith"
__credits__ = ["Jason Smith", "William Woods"]
__email__ = "jasons2@cisco.com"

"""
Base template for Python Scripts.
Created by Jason Smith (jasons2@cisco.com)  
"""

### TEMPLATE IMPORTS
import logging
import logging.config
import yaml
import argparse
import datetime

### IMPORTS
import requests
from openpyxl import load_workbook
import json
import xlsxwriter

## TEMPLATE FUNCTIONS
def setupLogging(logconfig, logfile, verbose, debug):
    """
    Setup logging configuration
    """
    global config

    with open(logconfig, 'rt') as f:
        config = yaml.safe_load(f.read())

    if logfile:
        configureLogging(handler='file_handler', filename=logfile)

    if verbose:
        configureLogging(handler='file_handler', formatter='verbose')
        configureLogging(handler='console', formatter='verbose')

    if debug:
        configureLogging(handler='file_handler', level='DEBUG')
        configureLogging(handler='console', level='DEBUG')

    logging.config.dictConfig(config)

def configureLogging(handler='', formatter='', level='', filename=''):
    """
    """
    global config, logger

    if formatter:
        config['handlers'][handler]['formatter'] = formatter
    if level:
        config['handlers'][handler]['level'] = level

    if filename and handler =='file_handler':
        config['handlers'][handler]['filename'] = filename

    logging.config.dictConfig(config)

def getArgs():
    """
    Get Arguments from Command Line.  Set defaults for logfile, logconfig, debug and verbose.
    """
    parser = argparse.ArgumentParser()
    parser.set_defaults(logfile='',
                        logconfig = 'logging.yaml',
                        debug=False,
                        verbose=False)

    # Default flags for arguments with defaults set.
    parser.add_argument('-l', '--logfile', help='Logging Filename (default is script_name.log)')
    parser.add_argument('-c', '--logconfig', help='Logging Configuration File (default is logging.yaml)')
    parser.add_argument('-d', '--debug', help='Set Debug', dest='debug', action='store_true')
    parser.add_argument('-v', '--verbose', help='Set Logging to Verbose', dest='verbose', action='store_true')
    parser.add_argument('-o', '--orginfo', help='Org Id or Org Name')
    parser.add_argument('-a', '--apikey', help='Set API Key', required=True)
    parser.add_argument('-e', '--excelfile', help='Set Input File', required=True)
    parser.add_argument('-f', '--filename', help='Output Filename', required=True)

    # Add additional arguments if required
    #parser.add_argument('-e', '--example', help='This is an example')

    output = parser.parse_args()  # assign contents of parser to output.

    if not output.logfile:  # Check to see if there is a logfile in command line.
        # Assign logfile to name of script plus .log
        output.logfile = "_" + __file__.split(".")[0] + ".log"

    return output

def info(msg):
    logger.info(msg)

def error(msg):
    logger.error(msg)

def warning(msg):
    logger.warning(msg)

def debug(msg):
    logger.debug(msg)

### HELPER FUNCTIONS
def getOrganizations(f_key):
    info("Gathering Oganizations")
    url = 'https://dashboard.meraki.com/api/v0/organizations'
    try:
        r = requests.get(url, headers={'X-Cisco-Meraki-API-Key': f_key, 'Content-Type': 'application/json'})

        if r.status_code != requests.codes.ok:
            return 'null'

        return r.json()

    except Exception as e:
        error(e)

def getOrganizationId(f_key, f_orginput):
    info("Determining Oganization ID")
    url = 'https://dashboard.meraki.com/api/v0/organizations'
    try:
        r = requests.get(url, headers={'X-Cisco-Meraki-API-Key': f_key, 'Content-Type': 'application/json'})

        if r.status_code != requests.codes.ok:
            return 'null'

        rjson = r.json()

        try:
            for record in rjson:
                if record['id'] == int(f_orginput):
                  return record['id']
        except:
            for record in rjson:
                if record['name'] == f_orginput:
                  return record['id']

        return('null')

    except Exception as e:
        error(e)

def getOrganizationName(f_key, f_orginput):
    info("Determining Oganization Name")
    url = 'https://dashboard.meraki.com/api/v0/organizations'
    try:
        r = requests.get(url, headers={'X-Cisco-Meraki-API-Key': f_key, 'Content-Type': 'application/json'})

        if r.status_code != requests.codes.ok:
            return 'null'

        rjson = r.json()

        try:
            for record in rjson:
                if record['id'] == int(f_orginput):
                  return record['name']
        except:
            for record in rjson:
                if record['name'] == f_orginput:
                  return record['name']

        return('null')

    except Exception as e:
        error(e)

def getShardURL(f_key, f_orgid):
    info("Gathering Shard URL")
    try:
        url = 'https://dashboard.meraki.com/api/v0/organizations/%s/snmp' % f_orgid
        r = requests.get(url, headers={'X-Cisco-Meraki-API-Key': f_key, 'Content-Type': 'application/json'})

        if r.status_code != requests.codes.ok:
            return 'null'

        rjson = r.json()

        return(rjson['hostname'])
    except Exception as e:
        error(e)

def getDeviceInfo(_fname, f_orgId):
    wb = load_workbook(filename=_fname, read_only=True)
    ws = wb['Networks']
    
    output = []
    
    for row in range(2, ws.max_row + 1):
        output.append({'name'     : ws.cell(row=row, column=1).value,
                       'timeZone' : ws.cell(row=row, column=2).value,
                       'tags'     : ws.cell(row=row, column=3).value,
                       'type'     : ws.cell(row=row, column=4).value})

    return output

def creatNetworks(f_apikey, f_url, f_org, f_networksToCreate):

    url = "https://%s/api/v0/organizations/%s/networks" % (f_url, f_org)

    payload = json.dumps(f_networksToCreate)

    headers = {
               'X-Cisco-Meraki-API-Key': f_apikey,
               'Content-Type': "application/json",
              }
    
    r = requests.request("POST", url, data=payload, headers=headers)
    
    return (r.json())

def createXL(f_fname, f_networks):
    wb = xlsxwriter.Workbook(f_fname)

    # Create Devices Tab
    info("Creating and Populating Device Tab")
    networkWS = wb.add_worksheet('Networks')
    networkColTitles = ['Net ID',
                        'Org ID',
                        'Type',
                        'Name',
                        'Time Zone',
                        'Tags',
                        'Disable My Meraki Com',
                        'Disable Remote Status Page'
                       ]
    networkWS.write_row(0,0,networkColTitles)
    row = 1
    for _net in f_networks:
        try:
            networkWS.write_row(row,0,_net)
        except Exception as e:
            error(e)
        row += 1
    info("Network Tab Complete")

### MAIN FUNCTION
def main():
    """
    Main Function
    """
    info("== (" + scriptStartTime + ") START SCRIPT ==")
    
    # Set Arguments    
    args = getArgs()
    
    # Assign API key from commmand line arguments
    apikey = args.apikey

    # Assign Excel Filename to inputFile variable
    inputFile = args.excelfile

    # Assign Filename from inout 
    filename = args.filename
    # Get Organization ID based on information provided by user
    # User gives either OrgID or OrgName and the program finds
    # OrgID to use for other API calls
    if args.orginfo:
        orgid = getOrganizationId(apikey, args.orginfo) # Get Organization ID
        orgName = getOrganizationName(apikey, args.orginfo) # Get Organization Name
    else:
        info("Must provide either Org Id or Organization Name.")
        info("Please use the list below and specify on command line with -o")

        # Get List of Organizations that the user has access to via the API
        for _org in getOrganizations(apikey):
            info("Name \"%s\"  Id \"%s\" " % (_org['name'], _org['id']))

        scriptEndTime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        info("== (" + scriptEndTime + ") END SCRIPT ==")
        exit()
    
    # Get ShardURL.  This is a requirement in Meraki to ensure API calls are 
    # made to the right shard (database server instance).
    shardurl = getShardURL(apikey, orgid)
    debug ("getShardURL returned %s" % shardurl)
    
    networksCreated = []
    networksToAdd = getDeviceInfo(inputFile, orgid)
    for _network in networksToAdd:
        response = creatNetworks(apikey, shardurl, orgid, _network)
        
        networksCreated.append([response['id'],
                                response['organizationId'],
                                response['type'],
                                response['name'],
                                response['timeZone'],
                                response['tags'],
                                response['disableMyMerakiCom'],
                                response['disableRemoteStatusPage']]
                              )
        
    createXL(filename, networksCreated)

    scriptEndTime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    info("== (" + scriptEndTime + ") END SCRIPT ==")

### MAIN PROGRAM
if __name__ == "__main__":
    # Record start time of script (after parser and compiler)
    scriptStartTime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Create initial logger. It will only log to console and will be disabled
    # when we read the logging configuration from the config file.
    logging.basicConfig(level="INFO",
                        format='%(asctime)s [%(levelname)s] : %(name)s : %(message)s',
                        datefmt='%D %I:%H:%M')
    logger = logging.getLogger()

    # If you want to configure logging for External Modules
    # Example of Configuring Logging Levels for External Modules
    # mlogger = logging.getLogger('mimir')
    # mlogger.setLevel(logging.INFO)

    # Get arguments from command line.
    args = getArgs()

    # Setup default logging based on default config file.  Previous Logging config is overwritten
    setupLogging(logconfig=args.logconfig, logfile=args.logfile, verbose=args.verbose, debug=args.debug)

    # Output arguments to log
    debug("==ARGUMENTS==")
    debug("logfile : %s" % (args.logfile))
    debug("logconfig: %s" % (args.logconfig))
    debug("debug: %s" % (args.debug))
    debug("verbose: %s" % (args.verbose))
    debug("==END ARGUMENTS==")
    # Add additional info if you add additional arguments above
    # debug("example: %s" % (args.example))

    main()
